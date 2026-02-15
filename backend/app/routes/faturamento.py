from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from ..database import get_db
from ..models.faturamento import Faturamento as FaturamentoModel
from ..models.projeto import Projeto as ProjetoModel
from ..models.funcionario import Funcionario as FuncionarioModel
from ..schemas.faturamento import FaturamentoCreate, FaturamentoUpdate, Faturamento
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

router = APIRouter()

# Rotas específicas PRIMEIRO (antes das genéricas com {id})

@router.get("/export/excel")
def exportar_excel(db: Session = Depends(get_db)):
    try:
        fats = db.query(FaturamentoModel).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Faturamentos"

        headers = ["ID", "Projeto ID", "Técnico ID", "Valor Faturado", "Data Faturamento", "Observações", "Data Criação", "Última Atualização"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for f in fats:
            ws.append([
                f.id,
                f.projeto_id,
                f.tecnico_id,
                float(f.valor_faturado) if f.valor_faturado is not None else 0,
                f.data_faturamento.isoformat() if f.data_faturamento else "",
                f.observacoes or "",
                f.criado_em.isoformat() if f.criado_em else "",
                f.atualizado_em.isoformat() if f.atualizado_em else "",
            ])

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=faturamentos.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


@router.post("/import/excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser em formato Excel (.xlsx ou .xls)")
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        # Expected minimal columns: Projeto ID, Técnico ID, Valor Faturado
        lower_headers = [ (h or "").lower() for h in headers ]
        required = {"projeto id", "técnico id", "tecnico id", "valor faturado"}
        if not any(r in lower_headers for r in ["projeto id","projeto_id"]) or not any(r in lower_headers for r in ["valor faturado","valor_faturado"]):
            raise HTTPException(status_code=400, detail="Arquivo inválido. Colunas obrigatórias: Projeto ID, Valor Faturado")

        erros = []
        inseridos = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Map by position: ID, Projeto ID, Técnico ID, Valor Faturado, Data Faturamento, Observações
                projeto_id = row[1]
                tecnico_id = row[2]
                valor = row[3] or 0
                data_fat = row[4]
                obs = row[5]

                if not projeto_id:
                    erros.append(f"Linha {idx}: Projeto ID ausente")
                    continue

                if not tecnico_id:
                    erros.append(f"Linha {idx}: Técnico ID ausente (obrigatório)")
                    continue

                projeto = db.query(ProjetoModel).filter(ProjetoModel.id == int(projeto_id)).first()
                if not projeto:
                    erros.append(f"Linha {idx}: Projeto {projeto_id} não encontrado")
                    continue

                func = db.query(FuncionarioModel).filter(FuncionarioModel.id == int(tecnico_id)).first()
                if not func:
                    erros.append(f"Linha {idx}: Técnico {tecnico_id} não encontrado")
                    continue

                novo = FaturamentoModel(
                    projeto_id=int(projeto_id),
                    tecnico_id=int(tecnico_id),
                    valor_faturado=float(valor),
                    data_faturamento=data_fat if isinstance(data_fat, datetime) else None,
                    observacoes=str(obs) if obs else None
                )
                db.add(novo)
                db.commit()
                db.refresh(novo)
                inseridos += 1
            except Exception as e:
                erros.append(f"Linha {idx}: Erro {str(e)}")

        return {"mensagem": f"Importação concluída. Inseridos: {inseridos}", "erros": erros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar: {str(e)}")


@router.get("/projeto/{projeto_id}", response_model=List[Faturamento])
def listar_por_projeto(projeto_id: int, db: Session = Depends(get_db)):
    return db.query(FaturamentoModel).filter(FaturamentoModel.projeto_id == projeto_id).all()


# Rotas genéricas (com {id}) DEPOIS

@router.post("/", response_model=Faturamento)
def criar_faturamento(fat: FaturamentoCreate, db: Session = Depends(get_db)):
    # Validate project exists
    projeto = db.query(ProjetoModel).filter(ProjetoModel.id == fat.projeto_id).first()
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    funcionario = db.query(FuncionarioModel).filter(FuncionarioModel.id == fat.tecnico_id).first()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionário (técnico) não encontrado")

    novo = FaturamentoModel(
        projeto_id=fat.projeto_id,
        tecnico_id=fat.tecnico_id,
        valor_faturado=fat.valor_faturado,
        data_faturamento=fat.data_faturamento,
        observacoes=fat.observacoes
    )
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

@router.get("/", response_model=List[Faturamento])
def listar_faturamentos(db: Session = Depends(get_db)):
    return db.query(FaturamentoModel).all()

@router.get("/{id}", response_model=Faturamento)
def obter_faturamento(id: int, db: Session = Depends(get_db)):
    fat = db.query(FaturamentoModel).filter(FaturamentoModel.id == id).first()
    if not fat:
        raise HTTPException(status_code=404, detail="Faturamento não encontrado")
    return fat

@router.put("/{id}", response_model=Faturamento)
def atualizar_faturamento(id: int, fat_update: FaturamentoUpdate, db: Session = Depends(get_db)):
    fat = db.query(FaturamentoModel).filter(FaturamentoModel.id == id).first()
    if not fat:
        raise HTTPException(status_code=404, detail="Faturamento não encontrado")
    for key, value in fat_update.dict(exclude_unset=True).items():
        setattr(fat, key, value)
    db.add(fat)
    db.commit()
    db.refresh(fat)
    return fat

@router.delete("/{id}")
def deletar_faturamento(id: int, db: Session = Depends(get_db)):
    fat = db.query(FaturamentoModel).filter(FaturamentoModel.id == id).first()
    if not fat:
        raise HTTPException(status_code=404, detail="Faturamento não encontrado")
    db.delete(fat)
    db.commit()
    return {"message": "Faturamento deletado"}

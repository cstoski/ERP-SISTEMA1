from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from io import BytesIO

from ..database import get_db
from ..models.funcionario import Funcionario as FuncionarioModel
from ..schemas.funcionario import Funcionario, FuncionarioCreate, FuncionarioUpdate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()


# Rotas com paths específicos
@router.get("/tecnicos")
def listar_tecnicos(db: Session = Depends(get_db)):
    """Listar todos os funcionários para seleção como técnico"""
    funcionarios = db.query(FuncionarioModel).all()
    return [
        {
            "id": f.id,
            "nome": f.nome,
            "departamento": f.departamento,
            "email": f.email
        }
        for f in funcionarios
    ]


@router.get("/export/excel")
def exportar_excel(db: Session = Depends(get_db)):
    """Exportar todos os funcionários para arquivo Excel"""
    try:
        funcionarios = db.query(FuncionarioModel).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Funcionários"
        
        # Definir headers
        headers = [
            "ID",
            "Nome",
            "Departamento",
            "Telefone Fixo",
            "Celular",
            "Email",
            "Data de Criação",
            "Última Atualização"
        ]
        
        ws.append(headers)
        
        # Estilo do header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for funcionario in funcionarios:
            ws.append([
                funcionario.id,
                funcionario.nome,
                funcionario.departamento or "",
                funcionario.telefone_fixo or "",
                funcionario.celular or "",
                funcionario.email or "",
                funcionario.criado_em.isoformat() if funcionario.criado_em else "",
                funcionario.atualizado_em.isoformat() if funcionario.atualizado_em else ""
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=funcionarios.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


@router.post("/import/excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar funcionários de arquivo Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser em formato Excel (.xlsx ou .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        funcionarios_importados = []
        erros = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                dados = {}
                headers_arquivo = [cell.value for cell in ws[1]]
                
                for header_idx, header in enumerate(headers_arquivo):
                    if header and row[header_idx] is not None:
                        header_lower = header.lower()
                        if header_lower == "id":
                            continue
                        elif header_lower == "nome":
                            dados['nome'] = row[header_idx]
                        elif header_lower == "departamento":
                            dados['departamento'] = row[header_idx]
                        elif header_lower == "telefone fixo":
                            dados['telefone_fixo'] = row[header_idx]
                        elif header_lower == "celular":
                            dados['celular'] = row[header_idx]
                        elif header_lower == "email":
                            dados['email'] = row[header_idx]
                
                # Validação obrigatória
                if not dados.get('nome'):
                    erros.append(f"Linha {idx}: Nome é obrigatório")
                    continue
                
                # Validar se funcionário com mesmo nome já existe
                funcionario_existente = db.query(FuncionarioModel).filter(
                    FuncionarioModel.nome == dados['nome']
                ).first()
                
                if funcionario_existente:
                    erros.append(f"Linha {idx}: Funcionário '{dados['nome']}' já existe")
                    continue
                
                novo_funcionario = FuncionarioModel(**dados)
                db.add(novo_funcionario)
                funcionarios_importados.append(dados.get('nome'))
                
            except Exception as e:
                if not any(f"Linha {idx}" in erro for erro in erros):
                    erros.append(f"Linha {idx}: {str(e)}")
        
        if funcionarios_importados:
            db.commit()
        
        return {
            "message": f"{len(funcionarios_importados)} funcionário(s) importado(s) com sucesso",
            "importados": funcionarios_importados,
            "erros": erros if erros else None,
            "total_sucesso": len(funcionarios_importados),
            "total_erros": len(erros)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar arquivo: {str(e)}")


# Rotas com paths dinâmicos
@router.post("/", response_model=Funcionario)
def criar_funcionario(funcionario: FuncionarioCreate, db: Session = Depends(get_db)):
    db_funcionario = FuncionarioModel(**funcionario.model_dump())
    db.add(db_funcionario)
    db.commit()
    db.refresh(db_funcionario)
    return db_funcionario


@router.get("/", response_model=List[Funcionario])
def ler_todos_funcionarios(db: Session = Depends(get_db)):
    print("DEBUG: GET /funcionarios/ foi chamado")
    funcionarios = db.query(FuncionarioModel).all()
    print(f"DEBUG: Retornando {len(funcionarios)} funcionários")
    return funcionarios


@router.get("/{funcionario_id}", response_model=Funcionario)
def obter_funcionario(funcionario_id: int, db: Session = Depends(get_db)):
    funcionario = db.query(FuncionarioModel).filter(FuncionarioModel.id == funcionario_id).first()
    if not funcionario:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    return funcionario


@router.put("/{funcionario_id}", response_model=Funcionario)
def atualizar_funcionario(funcionario_id: int, funcionario: FuncionarioUpdate, db: Session = Depends(get_db)):
    db_funcionario = db.query(FuncionarioModel).filter(FuncionarioModel.id == funcionario_id).first()
    if not db_funcionario:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    
    update_data = funcionario.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_funcionario, key, value)
        
    db.add(db_funcionario)
    db.commit()
    db.refresh(db_funcionario)
    return db_funcionario


@router.delete("/{funcionario_id}", response_model=Funcionario)
def deletar_funcionario(funcionario_id: int, db: Session = Depends(get_db)):
    db_funcionario = db.query(FuncionarioModel).filter(FuncionarioModel.id == funcionario_id).first()
    if not db_funcionario:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
        
    db.delete(db_funcionario)
    db.commit()
    return db_funcionario

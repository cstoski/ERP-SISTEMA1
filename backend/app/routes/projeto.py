from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from typing import List
from datetime import datetime
import os
from decimal import Decimal
from io import BytesIO

from ..database import get_db
from ..models.projeto import Projeto as ProjetoModel, StatusProjeto
from ..models.pessoa_juridica import PessoaJuridica as PessoaJuridicaModel
from ..models.contato import Contato as ContatoModel
from ..models.faturamento import Faturamento as FaturamentoModel
from ..models.user import User as UserModel
from ..config import settings
from ..schemas.projeto import Projeto, ProjetoCreate, ProjetoUpdate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from .auth import get_current_user

router = APIRouter()


def gerar_numero_projeto(db: Session, numero_manual: str = None) -> str:
    """Gera número de projeto automático ou valida manual"""
    if numero_manual:
        # Validar se número já existe
        existente = db.query(ProjetoModel).filter(ProjetoModel.numero == numero_manual).first()
        if existente:
            raise ValueError(f"Número de projeto '{numero_manual}' já existe")
        return numero_manual
    
    # Gerar automático: TC + ANO (2 dígitos) + MÊS (2 dígitos) + SEQUENCIAL (3 dígitos)
    now = datetime.now()
    ano = str(now.year)[-2:]  # Últimos 2 dígitos do ano
    mes = f"{now.month:02d}"   # Mês com zero à esquerda
    
    # Contar quantos projetos existem no ano atual com esse formato
    prefix = f"TC{ano}{mes}"
    
    # Contar projetos que começam com esse prefixo
    count = db.query(func.count(ProjetoModel.id)).filter(
        ProjetoModel.numero.like(f"{prefix}%")
    ).scalar() or 0
    
    sequencial = f"{count + 1:03d}"  # 3 dígitos com zeros à esquerda
    
    return f"{prefix}{sequencial}"


# Rotas com paths específicos
@router.get("/clientes", response_model=List[dict])
def listar_clientes(db: Session = Depends(get_db)):
    """Listar apenas pessoas jurídicas que são clientes"""
    clientes = db.query(PessoaJuridicaModel).filter(
        PessoaJuridicaModel.tipo == "Cliente"
    ).all()
    return [
        {
            "id": c.id,
            "razao_social": c.razao_social,
            "nome_fantasia": c.nome_fantasia,
            "cnpj": c.cnpj
        }
        for c in clientes
    ]


@router.get("/cliente/{cliente_id}/contatos", response_model=List[dict])
def listar_contatos_cliente(cliente_id: int, db: Session = Depends(get_db)):
    """Listar contatos de um cliente específico"""
    contatos = db.query(ContatoModel).filter(
        ContatoModel.pessoa_juridica_id == cliente_id
    ).all()
    return [
        {
            "id": c.id,
            "nome": c.nome,
            "email": c.email,
            "celular": c.celular
        }
        for c in contatos
    ]


@router.get("/proximo-numero")
def obter_proximo_numero(db: Session = Depends(get_db)):
    """Obter próximo número automático para projeto"""
    try:
        numero = gerar_numero_projeto(db)
        return {"numero": numero}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/export/excel")
def exportar_excel(db: Session = Depends(get_db)):
    """Exportar todos os projetos para arquivo Excel"""
    try:
        projetos = db.query(ProjetoModel).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Projetos"
        
        # Definir headers
        headers = [
            "Número",
            "Cliente",
            "Nome do Projeto",
            "Contato",
            "Técnico",
            "Valor Orçado",
            "Valor de Venda",
            "Prazo (dias)",
            "Data Pedido Compra",
            "Status",
            "Data de Criação",
            "Última Atualização"
        ]
        
        ws.append(headers)
        
        # Estilo do header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for projeto in projetos:
            cliente = db.query(PessoaJuridicaModel).filter(
                PessoaJuridicaModel.id == projeto.cliente_id
            ).first()
            contato = db.query(ContatoModel).filter(
                ContatoModel.id == projeto.contato_id
            ).first()
            
            cliente_nome = cliente.razao_social if cliente else "N/A"
            contato_nome = contato.nome if contato else "N/A"
            
            ws.append([
                projeto.numero,
                cliente_nome,
                projeto.nome,
                contato_nome,
                projeto.tecnico,
                float(projeto.valor_orcado) if projeto.valor_orcado else 0.00,
                float(projeto.valor_venda) if projeto.valor_venda else 0.00,
                projeto.prazo_entrega_dias,
                projeto.data_pedido_compra.isoformat() if projeto.data_pedido_compra else "",
                projeto.status.value,
                projeto.criado_em.isoformat() if projeto.criado_em else "",
                projeto.atualizado_em.isoformat() if projeto.atualizado_em else ""
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=projetos.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


def _format_currency_br(value: Decimal) -> str:
    try:
        val = float(value)
    except Exception:
        val = 0.0
    formatted = f"{val:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def _format_date_br(value: datetime) -> str:
    if not value:
        return "N/A"
    return value.strftime("%d/%m/%Y")


def _format_datetime_br(value: datetime) -> str:
    if not value:
        return "N/A"
    return value.strftime("%d/%m/%Y %H:%M")


@router.post("/import/excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar projetos de arquivo Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser em formato Excel (.xlsx ou .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        projetos_importados = []
        erros = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                dados = {}
                headers_arquivo = [cell.value for cell in ws[1]]
                
                for header_idx, header in enumerate(headers_arquivo):
                    if header and row[header_idx] is not None:
                        header_lower = header.lower()
                        if header_lower == "número":
                            dados['numero'] = str(row[header_idx])
                        elif header_lower == "cliente":
                            cliente = db.query(PessoaJuridicaModel).filter(
                                PessoaJuridicaModel.razao_social == row[header_idx]
                            ).first()
                            if not cliente:
                                erros.append(f"Linha {idx}: Cliente '{row[header_idx]}' não encontrado")
                                raise ValueError("Cliente não encontrado")
                            dados['cliente_id'] = cliente.id
                        elif header_lower == "nome do projeto":
                            dados['nome'] = row[header_idx]
                        elif header_lower == "contato":
                            contato = db.query(ContatoModel).filter(
                                ContatoModel.nome == row[header_idx]
                            ).first()
                            if not contato:
                                erros.append(f"Linha {idx}: Contato '{row[header_idx]}' não encontrado")
                                raise ValueError("Contato não encontrado")
                            dados['contato_id'] = contato.id
                        elif header_lower == "técnico":
                            dados['tecnico'] = row[header_idx]
                        elif header_lower == "valor orçado":
                            dados['valor_orcado'] = Decimal(str(row[header_idx]))
                        elif header_lower == "valor de venda":
                            dados['valor_venda'] = Decimal(str(row[header_idx]))
                        elif header_lower == "prazo (dias)":
                            dados['prazo_entrega_dias'] = int(row[header_idx])
                        elif header_lower == "data pedido compra" and row[header_idx]:
                            if isinstance(row[header_idx], datetime):
                                dados['data_pedido_compra'] = row[header_idx]
                        elif header_lower == "status":
                            dados['status'] = row[header_idx]
                
                # Validações obrigatórias
                if not dados.get('numero'):
                    erros.append(f"Linha {idx}: Número é obrigatório")
                    continue
                
                if not dados.get('cliente_id'):
                    erros.append(f"Linha {idx}: Cliente é obrigatório")
                    continue
                
                if not dados.get('nome'):
                    erros.append(f"Linha {idx}: Nome do projeto é obrigatório")
                    continue
                
                if not dados.get('contato_id'):
                    erros.append(f"Linha {idx}: Contato é obrigatório")
                    continue
                
                if not dados.get('tecnico'):
                    erros.append(f"Linha {idx}: Técnico é obrigatório")
                    continue
                
                # Validar duplicidade de número
                existente = db.query(ProjetoModel).filter(
                    ProjetoModel.numero == dados['numero']
                ).first()
                
                if existente:
                    erros.append(f"Linha {idx}: Projeto com número '{dados['numero']}' já existe")
                    continue
                
                novo_projeto = ProjetoModel(**dados)
                db.add(novo_projeto)
                projetos_importados.append(dados.get('numero'))
                
            except Exception as e:
                if not any(f"Linha {idx}" in erro for erro in erros):
                    erros.append(f"Linha {idx}: {str(e)}")
        
        if projetos_importados:
            db.commit()
        
        return {
            "message": f"{len(projetos_importados)} projeto(s) importado(s) com sucesso",
            "importados": projetos_importados,
            "erros": erros if erros else None,
            "total_sucesso": len(projetos_importados),
            "total_erros": len(erros)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar arquivo: {str(e)}")


@router.get("/{projeto_id}/export/pdf")
def exportar_pdf(
    projeto_id: int,
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Exportar projeto em PDF"""
    projeto = db.query(ProjetoModel).filter(ProjetoModel.id == projeto_id).first()
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")

    cliente = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.id == projeto.cliente_id).first()
    contato = db.query(ContatoModel).filter(ContatoModel.id == projeto.contato_id).first()

    cliente_nome = "N/A"
    if cliente:
        cliente_nome = cliente.nome_fantasia or cliente.razao_social

    contato_nome = contato.nome if contato else "N/A"
    status_valor = getattr(projeto.status, "value", projeto.status)
    faturamentos = db.query(FaturamentoModel).filter(FaturamentoModel.projeto_id == projeto.id).all()
    total_faturado = sum([float(f.valor_faturado or 0) for f in faturamentos])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=88, bottomMargin=52)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    subtitle_style = ParagraphStyle(
        name="Subtitle",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#1f2937"),
        spaceAfter=6,
    )
    label_style = ParagraphStyle(
        name="Label",
        parent=styles["BodyText"],
        fontSize=9,
        textColor=colors.HexColor("#6b7280"),
    )

    story = []
    story.append(Paragraph(f"Relatorio do Projeto {projeto.numero}", title_style))
    story.append(Spacer(1, 12))

    def add_section(title: str):
        story.append(Paragraph(title, subtitle_style))

    def add_table(rows: list):
        table = Table(rows, colWidths=[140, 360])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 12))

    add_section("Dados do Projeto")
    add_table(
        [
            ["Campo", "Valor"],
            ["Nome", projeto.nome or "N/A"],
            ["Status", str(status_valor)],
            ["Tecnico", projeto.tecnico or "N/A"],
            ["Prazo (dias)", str(projeto.prazo_entrega_dias or "N/A")],
            ["Data Pedido Compra", _format_date_br(projeto.data_pedido_compra)],
        ]
    )

    add_section("Cliente")
    add_table(
        [
            ["Campo", "Valor"],
            ["Nome", cliente_nome],
            ["CNPJ", cliente.cnpj if cliente else "N/A"],
            ["Cidade/UF", f"{cliente.cidade}/{cliente.estado}" if cliente and cliente.cidade and cliente.estado else "N/A"],
            ["Endereco", cliente.endereco if cliente else "N/A"],
        ]
    )

    add_section("Contato")
    contato_telefone = "N/A"
    if contato:
        contato_telefone = contato.celular or contato.telefone_fixo or "N/A"
    add_table(
        [
            ["Campo", "Valor"],
            ["Nome", contato_nome],
            ["Email", contato.email if contato else "N/A"],
            ["Telefone", contato_telefone],
        ]
    )

    add_section("Valores")
    add_table(
        [
            ["Campo", "Valor"],
            ["Valor Orcado", _format_currency_br(projeto.valor_orcado or 0)],
            ["Valor de Venda", _format_currency_br(projeto.valor_venda or 0)],
            ["Total Faturado", _format_currency_br(Decimal(str(total_faturado)))],
        ]
    )

    add_section("Historico")
    add_table(
        [
            ["Campo", "Valor"],
            ["Criado em", _format_datetime_br(projeto.criado_em)],
            ["Atualizado em", _format_datetime_br(projeto.atualizado_em)],
        ]
    )

    add_section("Faturamentos")
    if faturamentos:
        fat_rows = [["Data", "Tecnico", "Valor", "Observacoes"]]
        for fat in faturamentos:
            tecnico_nome = fat.tecnico.nome if fat.tecnico else "N/A"
            fat_rows.append(
                [
                    _format_date_br(fat.data_faturamento),
                    tecnico_nome,
                    _format_currency_br(fat.valor_faturado or 0),
                    fat.observacoes or "",
                ]
            )
        fat_table = Table(fat_rows, colWidths=[80, 140, 90, 190])
        fat_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(fat_table)
    else:
        story.append(Paragraph("Nenhum faturamento registrado para este projeto.", label_style))

    def draw_header_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()

        logo_path = settings.LOGO_PATH
        if logo_path and os.path.exists(logo_path):
            try:
                logo = ImageReader(logo_path)
                img_width, img_height = logo.getSize()
                max_width = 160
                max_height = 70
                scale = min(1.0, max_width / img_width, max_height / img_height)
                logo_width = img_width * scale
                logo_height = img_height * scale
                logo_x = 36
                canvas_obj.drawImage(
                    logo,
                    logo_x,
                    A4[1] - (logo_height + 20),
                    width=logo_width,
                    height=logo_height,
                    mask='auto'
                )
            except Exception:
                pass

        canvas_obj.setFont("Helvetica", 8.5)
        canvas_obj.setFillColor(colors.HexColor("#6b7280"))
        page_number = canvas_obj.getPageNumber()
        canvas_obj.drawRightString(A4[0] - 36, 24, f"Pagina {page_number}")
        canvas_obj.drawString(
            36,
            24,
            f"Gerado em: {_format_datetime_br(datetime.now())} | Usuario: {current_user.username}"
        )

        canvas_obj.restoreState()

    doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    buffer.seek(0)

    filename = f"projeto_{projeto.numero}.pdf"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Rotas com paths dinâmicos
@router.post("/", response_model=Projeto)
def criar_projeto(projeto: ProjetoCreate, db: Session = Depends(get_db)):
    try:
        # Validar/gerar número
        numero = gerar_numero_projeto(db, projeto.numero)
        
        db_projeto = ProjetoModel(
            **projeto.model_dump(exclude={'numero'}),
            numero=numero
        )
        db.add(db_projeto)
        db.commit()
        db.refresh(db_projeto)
        return db_projeto
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/", response_model=List[Projeto])
def ler_todos_projetos(db: Session = Depends(get_db)):
    projetos = db.query(ProjetoModel).all()
    return projetos


@router.get("/{projeto_id}", response_model=Projeto)
def obter_projeto(projeto_id: int, db: Session = Depends(get_db)):
    projeto = db.query(ProjetoModel).filter(ProjetoModel.id == projeto_id).first()
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    return projeto


@router.put("/{projeto_id}", response_model=Projeto)
def atualizar_projeto(projeto_id: int, projeto: ProjetoUpdate, db: Session = Depends(get_db)):
    db_projeto = db.query(ProjetoModel).filter(ProjetoModel.id == projeto_id).first()
    if not db_projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    update_data = projeto.model_dump(exclude_unset=True)
    
    # Se número foi alterado, validar unicidade
    if 'numero' in update_data and update_data['numero'] != db_projeto.numero:
        existente = db.query(ProjetoModel).filter(
            ProjetoModel.numero == update_data['numero']
        ).first()
        if existente:
            raise HTTPException(status_code=400, detail="Número de projeto já existe")
    
    for key, value in update_data.items():
        setattr(db_projeto, key, value)
    
    db.add(db_projeto)
    db.commit()
    db.refresh(db_projeto)
    return db_projeto


@router.delete("/{projeto_id}", response_model=Projeto)
def deletar_projeto(projeto_id: int, db: Session = Depends(get_db)):
    db_projeto = db.query(ProjetoModel).filter(ProjetoModel.id == projeto_id).first()
    if not db_projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    db.delete(db_projeto)
    db.commit()
    return db_projeto

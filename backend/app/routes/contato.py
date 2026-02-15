from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List

from ..database import get_db
from ..models.contato import Contato as ContatoModel
from ..models.pessoa_juridica import PessoaJuridica as PessoaJuridicaModel
from ..schemas.contato import Contato, ContatoCreate, ContatoUpdate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

router = APIRouter()


# Rotas com paths específicos (devem vir primeiro)
@router.get("/export/excel")
def exportar_excel(db: Session = Depends(get_db)):
    """Exportar todos os contatos para arquivo Excel"""
    try:
        contatos = db.query(ContatoModel).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contatos"
        
        # Definir headers
        headers = [
            "ID",
            "Nome",
            "Empresa",
            "Departamento",
            "Telefone Fixo",
            "Celular",
            "Email",
            "Data de Criação",
            "Última Atualização"
        ]
        
        ws.append(headers)
        
        # Estilo do header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for contato in contatos:
            pessoa = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.id == contato.pessoa_juridica_id).first()
            empresa_nome = pessoa.razao_social if pessoa else "N/A"
            
            ws.append([
                contato.id,
                contato.nome,
                empresa_nome,
                contato.departamento or "",
                contato.telefone_fixo or "",
                contato.celular or "",
                contato.email or "",
                contato.criado_em.isoformat() if contato.criado_em else "",
                contato.atualizado_em.isoformat() if contato.atualizado_em else ""
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=contatos.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


@router.post("/import/excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar contatos de arquivo Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser em formato Excel (.xlsx ou .xls)")
    
    try:
        # Ler arquivo
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Processar linhas
        contatos_importados = []
        erros = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Mapear colunas (case-insensitive)
                dados = {}
                headers_arquivo = [cell.value for cell in ws[1]]
                
                for header_idx, header in enumerate(headers_arquivo):
                    if header and row[header_idx] is not None:
                        header_lower = header.lower()
                        if header_lower == "id":
                            continue
                        elif header_lower == "nome":
                            dados['nome'] = row[header_idx]
                        elif header_lower == "empresa":
                            # Procurar empresa pelo nome
                            empresa = db.query(PessoaJuridicaModel).filter(
                                PessoaJuridicaModel.razao_social == row[header_idx]
                            ).first()
                            if not empresa:
                                erros.append(f"Linha {idx}: Empresa '{row[header_idx]}' não encontrada")
                                raise ValueError(f"Empresa não encontrada")
                            dados['pessoa_juridica_id'] = empresa.id
                        elif header_lower == "departamento":
                            dados['departamento'] = row[header_idx]
                        elif header_lower == "telefone fixo":
                            dados['telefone_fixo'] = row[header_idx]
                        elif header_lower == "celular":
                            dados['celular'] = row[header_idx]
                        elif header_lower == "email":
                            dados['email'] = row[header_idx]
                
                # Validações obrigatórias
                if not dados.get('nome'):
                    erros.append(f"Linha {idx}: Nome é obrigatório")
                    continue
                
                if not dados.get('pessoa_juridica_id'):
                    erros.append(f"Linha {idx}: Empresa é obrigatória")
                    continue
                
                # Validar se contato com mesmo nome e empresa já existe
                contato_existente = db.query(ContatoModel).filter(
                    ContatoModel.nome == dados['nome'],
                    ContatoModel.pessoa_juridica_id == dados['pessoa_juridica_id']
                ).first()
                
                if contato_existente:
                    erros.append(f"Linha {idx}: Contato '{dados['nome']}' já existe nessa empresa")
                    continue
                
                # Criar objeto
                novo_contato = ContatoModel(**dados)
                db.add(novo_contato)
                contatos_importados.append(dados.get('nome'))
                
            except Exception as e:
                if not any(f"Linha {idx}" in erro for erro in erros):
                    erros.append(f"Linha {idx}: {str(e)}")
        
        # Commit se houver contatos válidos
        if contatos_importados:
            db.commit()
        
        return {
            "message": f"{len(contatos_importados)} contato(s) importado(s) com sucesso",
            "importados": contatos_importados,
            "erros": erros if erros else None,
            "total_sucesso": len(contatos_importados),
            "total_erros": len(erros)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar arquivo: {str(e)}")


# Rotas com paths dinâmicos
@router.get("/pessoa/{pessoa_juridica_id}", response_model=List[Contato])
def ler_contatos_por_pessoa(pessoa_juridica_id: int, db: Session = Depends(get_db)):
    contatos = db.query(ContatoModel).filter(ContatoModel.pessoa_juridica_id == pessoa_juridica_id).all()
    return contatos

@router.post("/", response_model=Contato)
def criar_contato(contato: ContatoCreate, db: Session = Depends(get_db)):
    db_contato = ContatoModel(**contato.model_dump())
    db.add(db_contato)
    db.commit()
    db.refresh(db_contato)
    return db_contato

@router.get("/", response_model=List[Contato])
def ler_todos_contatos(db: Session = Depends(get_db)):
    contatos = db.query(ContatoModel).all()
    return contatos

@router.get("/{contato_id}", response_model=Contato)
def obter_contato(contato_id: int, db: Session = Depends(get_db)):
    contato = db.query(ContatoModel).filter(ContatoModel.id == contato_id).first()
    if not contato:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
    return contato

@router.put("/{contato_id}", response_model=Contato)
def atualizar_contato(contato_id: int, contato: ContatoUpdate, db: Session = Depends(get_db)):
    db_contato = db.query(ContatoModel).filter(ContatoModel.id == contato_id).first()
    if not db_contato:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
    
    update_data = contato.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_contato, key, value)
        
    db.add(db_contato)
    db.commit()
    db.refresh(db_contato)
    return db_contato

@router.delete("/{contato_id}", response_model=Contato)
def deletar_contato(contato_id: int, db: Session = Depends(get_db)):
    db_contato = db.query(ContatoModel).filter(ContatoModel.id == contato_id).first()
    if not db_contato:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
        
    db.delete(db_contato)
    db.commit()
    return db_contato
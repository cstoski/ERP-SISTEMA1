from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from ..database import get_db
from ..models.pessoa_juridica import PessoaJuridica as PessoaJuridicaModel
from ..schemas import pessoa_juridica as schemas
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os
from datetime import datetime

router = APIRouter()

@router.post("/", response_model=schemas.PessoaJuridica)
def criar_pessoa_juridica(pessoa: schemas.PessoaJuridicaCreate, db: Session = Depends(get_db)):
    db_cnpj = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.cnpj == pessoa.cnpj).first()
    if db_cnpj:
        raise HTTPException(status_code=400, detail="CNPJ já cadastrado")
        
    db_sigla = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.sigla == pessoa.sigla).first()
    if db_sigla:
        raise HTTPException(status_code=400, detail="Sigla já cadastrada")

    db_pessoa = PessoaJuridicaModel(**pessoa.model_dump())
    db.add(db_pessoa)
    db.commit()
    db.refresh(db_pessoa)
    return db_pessoa

@router.get("/", response_model=List[schemas.PessoaJuridica])
def listar_pessoas_juridicas(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    pessoas = db.query(PessoaJuridicaModel).offset(skip).limit(limit).all()
    return pessoas

@router.get("/{pessoa_id}", response_model=schemas.PessoaJuridica)
def obter_pessoa_juridica(pessoa_id: int, db: Session = Depends(get_db)):
    pessoa = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.id == pessoa_id).first()
    if pessoa is None:
        raise HTTPException(status_code=404, detail="Pessoa Jurídica não encontrada")
    return pessoa

@router.put("/{pessoa_id}", response_model=schemas.PessoaJuridica)
def atualizar_pessoa_juridica(pessoa_id: int, pessoa: schemas.PessoaJuridicaUpdate, db: Session = Depends(get_db)):
    db_pessoa = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.id == pessoa_id).first()
    if db_pessoa is None:
        raise HTTPException(status_code=404, detail="Pessoa Jurídica não encontrada")
    
    for key, value in pessoa.model_dump().items():
        setattr(db_pessoa, key, value)
    
    db.commit()
    db.refresh(db_pessoa)
    return db_pessoa

@router.delete("/{pessoa_id}")
def deletar_pessoa_juridica(pessoa_id: int, db: Session = Depends(get_db)):
    db_pessoa = db.query(PessoaJuridicaModel).filter(PessoaJuridicaModel.id == pessoa_id).first()
    if db_pessoa is None:
        raise HTTPException(status_code=404, detail="Pessoa Jurídica não encontrada")
    
    db.delete(db_pessoa)
    db.commit()
    return {"message": "Pessoa Jurídica deletada com sucesso"}


@router.get("/export/excel")
def exportar_excel(db: Session = Depends(get_db)):
    """Exportar todas as pessoas jurídicas para arquivo Excel"""
    try:
        pessoas = db.query(PessoaJuridicaModel).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pessoas Jurídicas"
        
        # Definir headers
        headers = [
            "ID",
            "Razão Social",
            "Nome Fantasia",
            "Sigla",
            "CNPJ",
            "Tipo",
            "Inscrição Estadual",
            "Inscrição Municipal",
            "Endereço",
            "Complemento",
            "Cidade",
            "Estado",
            "CEP",
            "País",
            "Data de Criação",
            "Última Atualização"
        ]
        
        ws.append(headers)
        
        # Estilo do header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for pessoa in pessoas:
            ws.append([
                pessoa.id,
                pessoa.razao_social,
                pessoa.nome_fantasia or "",
                pessoa.sigla,
                pessoa.cnpj,
                pessoa.tipo,
                pessoa.inscricao_estadual or "",
                pessoa.inscricao_municipal or "",
                pessoa.endereco or "",
                pessoa.complemento or "",
                pessoa.cidade or "",
                pessoa.estado or "",
                pessoa.cep or "",
                pessoa.pais or "",
                pessoa.criado_em.isoformat() if pessoa.criado_em else "",
                pessoa.atualizado_em.isoformat() if pessoa.atualizado_em else ""
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 20
        
        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=pessoas_juridicas.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")


@router.post("/import/excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar pessoas jurídicas de arquivo Excel
    
    Validações:
    - CNPJ não pode ser duplicado (incluindo duplicatas já cadastradas)
    - Sigla não pode ser duplicada
    - Arquivo deve ter o mesmo formato da exportação
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser em formato Excel (.xlsx ou .xls)")
    
    try:
        # Ler arquivo
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Verificar headers
        headers_esperados = [
            "Razão Social", "Nome Fantasia", "Sigla", "CNPJ", "Tipo",
            "Inscrição Estadual", "Inscrição Municipal", "Endereço", "Complemento",
            "Cidade", "Estado", "CEP", "País"
        ]
        
        headers_arquivo = [cell.value for cell in ws[1]]
        
        # Validar headers (flexível, buscando as colunas necessárias)
        colunas_requeridas = {"sigla", "cnpj", "razão social"}
        colunas_arquivo = {h.lower() if h else "" for h in headers_arquivo if h}
        
        if not colunas_requeridas.issubset(colunas_arquivo):
            raise HTTPException(
                status_code=400,
                detail="Arquivo inválido. Colunas obrigatórias: CNPJ, Sigla, Razão Social"
            )
        
        # Processar linhas
        pessoas_importadas = []
        erros = []
        cnpjs_arquivo = set()
        siglas_arquivo = set()
        
        # Obter CNPJs e siglas já cadastrados
        cnpjs_existentes = {p.cnpj for p in db.query(PessoaJuridicaModel.cnpj).all()}
        siglas_existentes = {p.sigla for p in db.query(PessoaJuridicaModel.sigla).all()}
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Mapear colunas (case-insensitive)
                dados = {}
                for header_idx, header in enumerate(headers_arquivo):
                    if header and row[header_idx] is not None:
                        header_lower = header.lower()
                        if header_lower == "id":
                            continue  # Ignorar ID
                        elif header_lower == "razão social":
                            dados['razao_social'] = row[header_idx]
                        elif header_lower == "nome fantasia":
                            dados['nome_fantasia'] = row[header_idx]
                        elif header_lower == "sigla":
                            dados['sigla'] = str(row[header_idx]).upper() if row[header_idx] else ""
                        elif header_lower == "cnpj":
                            dados['cnpj'] = str(row[header_idx]).replace('.', '').replace('/', '').replace('-', '')
                        elif header_lower == "tipo":
                            dados['tipo'] = row[header_idx] or "Cliente"
                        elif header_lower == "inscrição estadual":
                            dados['inscricao_estadual'] = row[header_idx]
                        elif header_lower == "inscrição municipal":
                            dados['inscricao_municipal'] = row[header_idx]
                        elif header_lower == "endereço":
                            dados['endereco'] = row[header_idx]
                        elif header_lower == "complemento":
                            dados['complemento'] = row[header_idx]
                        elif header_lower == "cidade":
                            dados['cidade'] = row[header_idx]
                        elif header_lower == "estado":
                            dados['estado'] = row[header_idx]
                        elif header_lower == "cep":
                            dados['cep'] = row[header_idx]
                        elif header_lower == "país":
                            dados['pais'] = row[header_idx]
                
                # Validações obrigatórias
                if not dados.get('razao_social'):
                    erros.append(f"Linha {idx}: Razão Social é obrigatória")
                    continue
                
                if not dados.get('cnpj'):
                    erros.append(f"Linha {idx}: CNPJ é obrigatório")
                    continue
                
                if not dados.get('sigla'):
                    erros.append(f"Linha {idx}: Sigla é obrigatória")
                    continue
                
                # Validar CNPJ duplicado na base existente
                if dados['cnpj'] in cnpjs_existentes:
                    erros.append(f"Linha {idx}: CNPJ {dados['cnpj']} já cadastrado no sistema")
                    continue
                
                # Validar CNPJ duplicado dentro do arquivo
                if dados['cnpj'] in cnpjs_arquivo:
                    erros.append(f"Linha {idx}: CNPJ {dados['cnpj']} duplicado no arquivo de importação")
                    continue
                
                # Validar sigla duplicada na base existente
                if dados['sigla'] in siglas_existentes:
                    erros.append(f"Linha {idx}: Sigla {dados['sigla']} já cadastrada no sistema")
                    continue
                
                # Validar sigla duplicada dentro do arquivo
                if dados['sigla'] in siglas_arquivo:
                    erros.append(f"Linha {idx}: Sigla {dados['sigla']} duplicada no arquivo de importação")
                    continue
                
                cnpjs_arquivo.add(dados['cnpj'])
                siglas_arquivo.add(dados['sigla'])
                
                # Criar objeto
                nova_pessoa = PessoaJuridicaModel(**dados)
                db.add(nova_pessoa)
                pessoas_importadas.append(dados.get('razao_social'))
                
            except Exception as e:
                erros.append(f"Linha {idx}: {str(e)}")
        
        # Commit se houver pessoas válidas
        if pessoas_importadas:
            db.commit()
        
        return {
            "message": f"{len(pessoas_importadas)} pessoa(s) jurídica(s) importada(s) com sucesso",
            "importadas": pessoas_importadas,
            "erros": erros if erros else None,
            "total_sucesso": len(pessoas_importadas),
            "total_erros": len(erros)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao importar arquivo: {str(e)}")